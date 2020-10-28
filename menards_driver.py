from product_websites.lowes_base import Lowes
from product_websites.menards_base import Menards
import os
from os import path
from datetime import datetime
from configparser import ConfigParser
from multiprocessing import Process
from openpyxl import load_workbook


def check_create_dir(dirname):
    '''
    Checks if directory exists and if it doesn't creates a new directory
    :param dirname: Path to directory
    '''
    if not path.exists(dirname):
        if '/' in dirname:
            os.makedirs(dirname)
        else:
            os.mkdir(dirname)


def process_menards(inputfile, wait_time, debug=False):
    wb = load_workbook(inputfile)
    ws = wb.active
    with Menards('dump', debug) as scrape_obj:
        for site in ws['A']:
            scrape_obj.run_scrape(site.value, wait_time)


if __name__ == '__main__':
    print('Product Scraper 1.0')
    print('Reading config...')
    # Init config
    config = ConfigParser()
    config.read('masterconfig.ini')
    menards_input = config['menards']['input']
    menards_wait = int(config['menards']['wait_time'])
    debug_mode = config.getboolean('debug', 'enable_debug')

    # Init logging
    log_dir = 'logs/{}'.format(datetime.now().strftime('%y-%m-%d'))
    check_create_dir(log_dir)
    check_create_dir('screenshots')

    print('Creating directories...')
    check_create_dir('dump')

    process_menards(menards_input, menards_wait, debug_mode)
