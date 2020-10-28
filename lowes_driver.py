from product_websites.lowes_base import Lowes
import os
from os import path
from datetime import datetime
from configparser import ConfigParser
from openpyxl import load_workbook


def check_create_dir(dirname):
    '''
    Checks if directory exists and if it doesn't creates a new directory
    :param dirname: Path to directory
    '''
    if not path.exists(dirname):
        if '/' in dirname:
            os.makedirs(dirname)
        else:
            os.mkdir(dirname)


def process_lowes(inputfile, wait_time, debug=False):
    wb = load_workbook(inputfile)
    ws = wb.active
    with Lowes('dump', debug) as scrape_obj:
        for site in ws['A'][1:]:
            scrape_obj.run_scrape(site.value, wait_time)


if __name__ == '__main__':
    print('Product Scraper 1.0')
    print('Reading config...')
    # Init config
    config = ConfigParser()
    config.read('masterconfig.ini')
    lowes_input = config['lowes']['input']
    lowes_wait = int(config['lowes']['wait_time'])
    debug_mode = config.getboolean('debug', 'enable_debug')

    # Init logging
    log_dir = 'logs/{}'.format(datetime.now().strftime('%y-%m-%d'))
    check_create_dir(log_dir)
    check_create_dir('screenshots')

    print('Creating directories...')
    check_create_dir('dump')

    process_lowes(lowes_input, lowes_wait, debug_mode)
