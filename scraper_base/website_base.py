from bs4 import BeautifulSoup
import os
from time import sleep
import json
from openpyxl import Workbook, load_workbook
import copy
import logging
from datetime import datetime
from os import path
from selenium.common.exceptions import ElementClickInterceptedException, StaleElementReferenceException
import sys
import undetected_chromedriver as uc
uc.install()
from selenium import webdriver


class WebsiteBase:
    '''
    This is a Framework for web scrapers. It opens and closes Selenium(Undetectable, Headless) automatically by the use
     of the with' keyword. It also facilitates scraping with various standard routines built-in including the main
     scraping and file saving routine. These can be overridden if desired.

    To make a new scraper, you need to inherit this class as a base class, update the 'scrape map' and override the
    'get_scrape_master' method as shown in the two subclasses in this project.

    The framework consults the scrape map, resolves the map into the required scrape methods and runs them to get the
    results. It then automatically calls the 'save_into_file' method which can be overridden (Saves into Excel
     by defualt)
    '''

    bin_dir = 'chrome-bin/chrome.exe'
    name = 'This_is_an_error'

    def __init__(self, out_path, debug_mode=False):
        self.outputpath = out_path + '-' + datetime.today().strftime('%Y-%m-%d') + '.xlsx'
        self.debug_mode = debug_mode
        self.results = []
        self.driver = None
        self.scrape_array = None
        self.scrape_map = None
        self.scrape_results = None
        self.scrape_obj = None
        self.log_dir = 'logs/{}'.format(datetime.now().strftime('%y-%m-%d'))
        self.logger = None
        self.init_logs()
        self.run_count = 1

    def __enter__(self):
        '''
        Automatic Selenium open
        '''
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1280x1696')
        chrome_options.add_argument('--user-data-dir={}'.format(os.getcwd() + '/chrome-data/user-data'))
        chrome_options.add_argument('--hide-scrollbars')
        chrome_options.add_argument("disable-infobars")
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument('--enable-logging')
        chrome_options.add_argument('--log-level=0')
        chrome_options.add_argument('--v=99')
        chrome_options.add_argument('--single-process')
        chrome_options.add_argument('--data-path={}'.format(os.getcwd() + '/chrome-data/data-path'))
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument("--incognito")
        chrome_options.add_argument("--disable-plugins-discovery")
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument('--homedir={}'.format(os.getcwd()))
        chrome_options.add_argument('--disk-cache-dir={}'.format(os.getcwd() + '/chrome-data/cache-dir'))
        chrome_options.add_argument(
            'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36')
        chrome_options.binary_location = WebsiteBase.bin_dir
        self.driver = webdriver.Chrome(options=chrome_options)
        return self

    def __exit__(self, tipe, value, traceback):
        '''
        Automatic Selemium close with error handling
        '''
        if tipe is not None:
            self.logger.error('Unexpected error:\n{0}\n{1}\n{2}'.format(tipe, value, traceback))
        self.driver.close()
        self.driver.quit()
    
    def init_logs(self):
        '''
        Initiates logging
        '''
        self.logger = logging.getLogger(self.name)
        self.logger.setLevel(logging.DEBUG)
        fileHandler = logging.FileHandler(path.join(self.log_dir, '{}.log'.format(self.name)))
        fileHandler.setFormatter(logging.Formatter('%(asctime)s:-[%(name)s] - %(levelname)s - %(message)s'))
        fileHandler.setLevel(logging.DEBUG)
        self.logger.addHandler(fileHandler)
        sys.stdout = open("logfile.txt", 'w')

    def get_scrape_master(self):
        '''
        Returns reference to object's bound scrape_methods object
        (MUST BE OVERRIDDEN IN DERIVED CLASSES)
        :return: reference to bound scrape_methods object
        '''
        return self.scrape_methods

    class scrape_methods:
        '''
        This inner class houses all the scrape methods. It contains some basic scrape methods but can be inherited on
        a derived class and have more specialized methods added to it. If new methods are added, do not forget to update
        the method map with new number codes pointing to your new methods
        '''
        def __init__(self, driver, logger, subclass_name):
            self.driver = driver
            self.logger = logger
            self.subclass_name = subclass_name
            # Default method mapping. DO NOT OVERRIDE, CONTINUE THIS NUMBERING SCHEME FOR NEW METHODS
            self.method_map = {1: self.simple_scrape, 2: self.simple_scrape_tag, 3: self.block_scrape,
                               4: self.block_scrape_tag}

        @classmethod
        def xpath_soup(cls, element):
            """
               Generate xpath from BeautifulSoup4 element.
               :param element: BeautifulSoup4 element.
               :type element: bs4.element.Tag or bs4.element.NavigableString
               :return: xpath as string
               """
            components = []
            child = element if element.name else element.parent
            for parent in child.parents:
                siblings = parent.find_all(child.name, recursive=False)
                components.append(
                    child.name if 1 == len(siblings) else '%s[%d]' % (
                        child.name,
                        next(i for i, s in enumerate(siblings, 1) if s is child)
                    )
                )
                child = parent
            components.reverse()
            return '/%s' % '/'.join(components)

        def get_class(self, keyword, soup, tag, element):
            '''
            Find class with keyword
            :param keyword: Keyword to look for
            :param soup: Reference to soup object
            :param tag: Tag to search
            :param element: Element to search
            :return: Soup object with tag that contains keyword in its class attribute
            '''
            for soup_ele in soup.find_all(tag):
                try:
                    target = soup_ele[element]
                    if keyword in target:
                        if isinstance(target, list):
                            return ' '.join(target)
                        else:
                            return target
                except KeyError:
                    continue
            return None

        def simple_scrape(self, soup, tag, element, element_value, target_element=None):
            '''
            Scrape text or attribute in a single tag
            :param soup: Reference to soup object
            :param tag: Tag to search
            :param element: Element to search
            :param element_value: Attribute to search
            :param target_element: Attribute to scrape
            :return: Scraped data
            '''
            if target_element is None:
                return soup.find_all(tag, {element: element_value})[0].text
            else:
                return soup.find_all(tag, {element: element_value})[0][target_element]

        def simple_scrape_tag(self, soup, tag, target_element=None):
            '''
            Simple scrape without checking for attribute value
            param soup: Reference to soup object
            :param tag: Tag to search
            :param target_element: Attribute to scrape
            :return: Scraped data
            '''
            if target_element is None:
                return soup.find_all(tag)[0].text
            else:
                return soup.find_all(tag)[0][target_element]

        def block_scrape(self, soup, tag, element, element_value, target_element=None, delimiter=','):
            '''
            Scrape a collection of elements
            :param soup: Reference to soup object
            :param tag: Target block tag
            :param element: Target block element
            :param element_value: Target attribute value
            :param target_element: Element to target inside of block
            :param delimiter: Character to use to separate each data scraped
            :return: Scraped data
            '''
            ret = []
            for tag_ele in soup.find_all(tag, {element: element_value}):
                if target_element is None:
                    ret.append(tag_ele.text)
                else:
                    ret.append(tag_ele[target_element])
            return delimiter.join(ret)

        def block_scrape_tag(self, soup, tag, target_element=None, delimiter=','):
            '''
            Scrape a collection of elements without checking for attribute value
            :param soup: Reference to soup object
            :param tag: Target block tag
            :param target_element: Target block element
            :param delimiter: Character to use to separate each data scraped
            :return: Scraped data
            '''
            ret = []
            for tag_ele in soup.find_all(tag):
                if target_element is None:
                    ret.append(tag_ele.text)
                else:
                    ret.append(tag_ele[target_element])
            return delimiter.join(ret)

    def save_into_file(self):
        '''
        Save data into file
        '''
        if os.path.isfile(self.outputpath):
            wb = load_workbook(self.outputpath)
            ws = wb.active
            next_row = len(ws['A']) + 1
            for index, value in enumerate(self.scrape_results.values()):
                ws.cell(row=next_row, column=index + 1, value=value)
            wb.save(self.outputpath)
        else:
            # Existing file with previous data not found. Assumed to be first time scraping
            wb = Workbook()
            ws = wb.active
            for index, (key, value) in enumerate(self.scrape_results.items()):
                # Filling column titles
                ws.cell(row=1, column=index + 1, value=key)
                # Filling first row
                ws.cell(row=2, column=index + 1, value=value)
            wb.save(self.outputpath)

    def refresh_scrape_array(self):
        '''
        The scrape map is converted into a dictionary that contains references to the appropriate scrape method (Python
        methods can be treated as objects if not called). At the end of each scrape run, the references are called and
        value in the dictionary changes to the called value.
        This problem is alleviated by creating a new object of the inner scrape_method class and building the
        scrape_array dictionary again with fresh references  
        '''
        self.logger.debug('Resfreshing function pointers')
        self.scrape_obj = self.get_scrape_master()(self.driver, self.logger, self.name)
        self.logger.debug('Creating new scrape array')
        new_scrape_map = copy.deepcopy(self.scrape_map)
        self.scrape_array = dict()
        for key, value in new_scrape_map.items():
            array_unit = value
            array_unit['method'] = self.scrape_obj.method_map[array_unit['method']]
            self.scrape_array[key] = array_unit

    def resolver(self, soup, array_unit):
        '''
        Resolves the method code into a refrence to the appropriate scrape method and calls it with the respective
         arguments
        :param soup:
        :param array_unit:
        :return: Result of the scrape method
        '''
        scrape_method = array_unit['method']
        array_unit.pop('method')
        array_unit['soup'] = soup
        result = scrape_method(**array_unit)
        if isinstance(result, dict):
            return self.convert_list(result)
        else:
            return result

    def convert_list(self, ret):
        filename = ret['filename']
        target = ret['data']
        wb = Workbook()
        ws = wb.active
        for index, key in enumerate(target[0].keys()):
            ws.cell(row=1, column=index + 1, value=key)
        for tar_index, target_dict in enumerate(target):
            for index, key in enumerate(target_dict.keys()):
                conv_ele = target_dict[key]
                self.logger.debug('Trying to convert: {}'.format(conv_ele))
                if isinstance(conv_ele, list):
                    if len(conv_ele) > 0:
                        if isinstance(conv_ele[0], dict):
                            conv_ele = json.dumps(conv_ele[0])
                        else:
                            conv_ele = ','.join(conv_ele)
                    else:
                        conv_ele = 'No data'
                ws.cell(row=tar_index + 2, column=index + 1, value=conv_ele)
        wb.save('dump/{}'.format(filename))
        return 'See {}'.format(filename)

    def refresh_cookies(self):
        self.driver.delete_all_cookies()

    def run_scrape(self, url, load_delay=3):
        '''
        Main scraping routine. This method takes a url as an input and loops through each entry of the scrape_array
        and runs them through the Resolver. (Can be overridden)
        :param url: String with url to be scraped
        :param load_delay: Optional time to wait before taking the page source after initial page load
        '''
        self.logger.debug('Scraping: {}'.format(url))
        self.scrape_results = dict()
        self.scrape_results['Url'] = url
        self.logger.info('Clearing cache')
        # Delete cookies before reconnecting to prevent detection
        self.refresh_cookies()
        self.driver.get(url)
        sleep(load_delay)
        if self.debug_mode:
            self.logger.debug('Saving screenshot')
            self.driver.save_screenshot(os.getcwd() + '/screenshots/{0}-{1}.png'.format(self.name, self.run_count))
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        self.refresh_scrape_array()
        for key, array_unit in self.scrape_array.items():
            try:
                self.scrape_results[key] = self.resolver(soup, array_unit)
            except IndexError:
                self.scrape_results[key] = 'N/A'
            except (ElementClickInterceptedException, StaleElementReferenceException):
                self.scrape_results[key] = 'N/A'
        self.run_count += 1
        self.save_into_file()

