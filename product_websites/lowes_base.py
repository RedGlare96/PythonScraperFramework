from bs4 import BeautifulSoup
from time import sleep
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, ElementClickInterceptedException
from .website_base import WebsiteBase


class Lowes(WebsiteBase):

    name = 'Lowes'

    def __init__(self, out_path, debug_mode=False):
        super(Lowes, self).__init__(out_path + '/Lowes', debug_mode)
        self.scrape_map = {'Price': {'method': 8},
                           'Model-no': {'method': 4, 'mode': 1},
                           'Item-no': {'method': 4, 'mode': 0},
                           'Section': {'method': 3, 'tag': 'li', 'element': 'class',
                                       'element_value': 'sc-fzpans liKQQk js-breadcrumb breadcrumb-item'},
                           'Image links': {'method': 7},
                           'Description': {'method': 1, 'tag': 'div', 'element': 'class',
                                           'element_value': 'romance'},
                           'Specs': {'method': 1, 'tag': 'div', 'element': 'class',
                                     'element_value': 'specs'},
                           'Reviews': {'method': 5},
                           'Questions and Answers': {'method': 6}}

    def get_scrape_master(self):
        return self.scrape_methods

    class scrape_methods(WebsiteBase.scrape_methods):
        def __init__(self, driver, logger, name):
            super(Lowes.scrape_methods, self).__init__(driver, logger, name)
            new_codes = {4: self.get_nos, 5: self.get_reviews, 6: self.get_qanda, 7: self.get_imgs, 8: self.get_price}
            self.method_map.update(new_codes)

        def get_price(self, soup):
            return soup.find_all('span', {'class': 'finalPrice'})[0].find_all('div')[0].text

        def get_nos(self, soup, mode=0):
            '''
                     Custom scrape method
            '''
            item_model = soup.find_all('div', {'class': 'modelNo'})[0]
            if mode == 0:
                return item_model.find_all('span')[0].text
            elif mode == 1:
                return item_model.find_all('span')[1].text
            else:
                return IndexError

        def get_reviews(self, soup):
            '''
                     Custom scrape method
            '''
            self.logger.debug('Start reviews')
            ret = []
            reviews_button_class2 = self.get_class('fsReviewLinkDefaultContainer', soup, 'div', 'class')
            reviews_button_tag = soup.find_all('div', {'class': reviews_button_class2})[0]
            reviews_button = self.driver.find_element_by_xpath(self.xpath_soup(reviews_button_tag))
            reviews_button.click()
            self.logger.debug('main button clicked')
            sleep(3)
            startreview_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            reviews_button_class = self.get_class('showMoreBtn', startreview_soup, 'button', 'class')
            self.logger.debug(reviews_button_class)
            reviews_loadmore_button = startreview_soup.find_all('button', {'class': reviews_button_class})[0]
            self.logger.debug('loadmore found')
            reviews_lodmore = self.xpath_soup(reviews_loadmore_button)
            click_count = 0
            while True:
                try:
                    self.logger.debug('Try {}'.format(click_count))
                    self.driver.execute_script("arguments[0].scrollIntoView();", reviews_button)
                    self.driver.find_element_by_xpath(reviews_lodmore).click()
                    click_count += 1
                    if click_count > 5:
                        break
                except NoSuchElementException:
                    break
                except StaleElementReferenceException:
                    break
                except ElementClickInterceptedException:
                    break
            review_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            self.logger.debug('Clicking readmore')
            readmore_class = self.get_class('padReset', startreview_soup, 'div', 'class')
            try:
                for readmore_ele in review_soup.find_all('div', {'class': readmore_class}):
                    try:
                        readmore_button = self.xpath_soup(readmore_ele.find_all('button')[0])
                        self.driver.find_element_by_xpath(readmore_button).click()
                    except IndexError:
                        continue
            except (ElementClickInterceptedException, NoSuchElementException):
                self.logger.debug('Could not find readmore')
            sleep(2)
            review_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            self.logger.debug('++++++++++reviews+++++++\n')
            review_class = self.get_class('reviewEach', review_soup, 'div', 'class')
            for i, review_ele in enumerate(review_soup.find_all('div', {'class': review_class})):
                try:
                    results = dict()
                    self.logger.debug('------Review {}------'.format(i + 1))
                    results['Reviewer and Date'] = review_ele['aria-label']
                    try:
                        results['Title'] = review_ele.find_all('h4', {'class': 'reviewEachTitle'})[0].text
                    except IndexError:
                        results['Title'] = 'N/A'
                    results['Rating'] = review_ele.find_all('div', {'class': 'reviewEachRating'})[0]['aria-label']
                    results['Content'] = review_ele.find_all('p', {'class': 'reviewRowContent'})[0].text
                    helpful_container = review_ele.find_all('div', {'class': 'helpfulContainer'})[0]
                    if helpful_container.contents:
                        buttons = helpful_container.find_all('button')
                        if len(buttons) == 4:
                            self.logger.debug('Responses found')
                            responses = []
                            comment_button = self.xpath_soup(buttons[2])
                            self.driver.find_element_by_xpath(comment_button).click()
                            comment_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
                            response_class = self.get_class('reviewEach', comment_soup, 'div', 'class')
                            for j, comment_ele in enumerate(
                                    comment_soup.find_all('div', {'class': response_class})[i].find_all('div', {
                                        'class': 'eachCommentBox'})):
                                response = dict()
                                self.logger.debug('-----Response {}-----'.format(j + 1))
                                response['Comment'] = comment_ele.find_all('p', {'class': 'commentParagraph'})[0].text
                                response['Commenter'] = comment_ele.find_all('p', {'class': 'commentHelper'})[0].text
                                responses.append(response)
                            results['Response'] = responses
                    ret.append(results)
                except (ElementClickInterceptedException, IndexError, KeyError):
                    continue
            if len(ret) > 0:
                return {'filename': 'lowes-reviews-{}.xlsx'.format(datetime.now().strftime('%Y-%m-%d')),
                        'keyword': 'reviews', 'data': ret}
            else:
                return 'No data'

        def get_qanda(self, soup):
            self.logger.debug('Start qanda')
            ret = []
            self.driver.refresh()
            sleep(2)
            startq_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            try:
                qanda_ele = startq_soup.find_all('button', {'class': 'qa fsQnaLinkDefault'})[0]
                qanda_button = self.xpath_soup(qanda_ele)
                self.driver.find_element_by_xpath(qanda_button).click()
            except (IndexError, ElementClickInterceptedException, StaleElementReferenceException):
                self.logger.debug('Could not click main button. Trying alternate')
                qanda_ele = startq_soup.find_all('div', {'aria-label': 'Community Q & A'})[0]
                qanda_button = self.xpath_soup(qanda_ele)
                self.driver.find_element_by_xpath(qanda_button).click()
            self.logger.debug('Main button clicked')
            sleep(2)
            self.driver.save_screenshot('qanda1.png')
            qanda_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            click_count = 0
            qanda_button_class = self.get_class('showMoreBtn', qanda_soup, 'button', 'class')
            self.logger.debug(qanda_button_class)
            qanda_loadmore = qanda_soup.find_all('button', {'class': qanda_button_class})[0]
            self.logger.debug('Loadmore button found')
            qanda_loadmore_button = self.xpath_soup(qanda_loadmore)
            while True:
                try:
                    self.logger.debug('Try {}'.format(click_count))
                    self.driver.find_element_by_xpath(qanda_loadmore_button).click()
                    click_count += 1
                    if click_count > 4:
                        break
                except NoSuchElementException:
                    break
                except StaleElementReferenceException:
                    break
                except ElementClickInterceptedException:
                    break
            self.logger.debug('++++++++++++++++++++ Q and A++++++++++++++++++++++++++++')
            question_class = self.get_class('questionEach', qanda_soup, 'div', 'class')
            for k, qanda_ele in enumerate(qanda_soup.find_all('div', {'class': question_class})):
                try:
                    results = dict()
                    self.logger.debug('-----Question {}-----'.format(k + 1))
                    results['Question Asker'] = qanda_ele['aria-label']
                    try:
                        results['Question'] = qanda_ele.find_all('span', {'class': 'questionTitle'})[0].text
                    except IndexError:
                        results['Question'] = qanda_ele.find_all('span', {'class': 'qnaSummary'})[0].text
                    try:
                        title_button = self.get_class('titleBtn', qanda_ele, 'button', 'class')
                        a_ele = qanda_ele.find_all('button', {'class': title_button})[0]
                        self.driver.find_element_by_xpath(self.xpath_soup(a_ele)).click()
                        sleep(1)
                        qanda_updated_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
                        self.logger.debug('---Answer---')
                        answers = []
                        for ans_ele in qanda_updated_soup.find_all('div', {'class': question_class})[k].find_all(
                                'div', {'class': 'answerRow'}):
                            answers.append(ans_ele.find_all('p')[0].text)
                        results['Answers'] = answers
                    except IndexError:
                        # Some entries have questions but no answers
                        results['Answers'] = 'N/A'
                    ret.append(results)
                except (IndexError, KeyError, NoSuchElementException):
                    continue
                except ElementClickInterceptedException:
                    continue
            if len(ret) > 0:
                return {'filename': 'lowes-qanda-{}.xlsx'.format(datetime.now().strftime('%Y-%m-%d')),
                        'keyword': 'qanda', 'data': ret}
            else:
                return 'No data'

        def get_imgs(self, soup):
            '''
                     Custom scrape method
            '''
            ret = []
            for imgtag in soup.find_all('li', {'data-testid': 'gal-thumbnail-item'}):
                for img_ele in imgtag.find_all('img'):
                    ret.append(img_ele['src'])
            return ','.join(ret)

    def convert_list(self, ret):
        filename = ret['filename']
        target = ret['data']
        keyword = ret['keyword']
        file_path = 'dump/' + filename
        insert_dict = {'Url': self.scrape_results['Url'], 'Model-no': self.scrape_results['Model-no'],
                       'Item-no': self.scrape_results['Item-no']}
        if os.path.isfile(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
            next_row = len(ws['A']) + 1
        else:
            model_keys = []
            wb = Workbook()
            ws = wb.active
            if keyword == 'reviews':
                model_keys = ['Url', 'Model-no', 'Item-no', 'Reviewer and Date', 'Title', 'Rating', 'Content', 'Comment', 'Commenter']
            if keyword == 'qanda':
                model_keys = ['Url', 'Model-no', 'Item-no', 'Question Asker', 'Question', 'Answers']
            for index, key in enumerate(model_keys):
                ws.cell(row=1, column=index + 1, value=key)
            next_row = 2
        for tar_ele in target:
            total_dict = {**insert_dict, **tar_ele}
            self.logger.debug('Total_dict: {}'.format(total_dict))
            for index, key in enumerate(total_dict.keys()):
                if keyword == 'reviews' and key == 'Response':
                    response = total_dict[key]
                    if isinstance(response, list):
                        ws.cell(row=next_row, column=index + 1, value=response[0]['Comment'])
                        ws.cell(row=next_row, column=index + 2, value=response[0]['Commenter'])
                        for val_ele in response[1:]:
                            next_row += 1
                            for inner_index, inner_key in enumerate(list(total_dict.keys())[:index]):
                                ws.cell(row=next_row, column=inner_index + 1, value=total_dict[inner_key])
                            ws.cell(row=next_row, column=index + 1, value=val_ele['Comment'])
                            ws.cell(row=next_row, column=index + 2, value=val_ele['Commenter'])
                    else:
                        ws.cell(row=next_row, column=index + 1, value=response['Comment'])
                        ws.cell(row=next_row, column=index + 2, value=response['Commenter'])
                elif keyword == 'qanda' and key == 'Answers':
                    answers = total_dict[key]
                    if isinstance(answers, list):
                        ws.cell(row=next_row, column=index + 1, value=answers[0])
                        for ans_ele in answers[1:]:
                            next_row += 1
                            for inner_index, inner_key in enumerate(list(total_dict.keys())[:index]):
                                ws.cell(row=next_row, column=inner_index + 1, value=total_dict[inner_key])
                            ws.cell(row=next_row, column=index + 1, value=ans_ele)
                    else:
                        ws.cell(row=next_row, column=index + 1, value=answers)
                else:
                    ws.cell(row=next_row, column=index + 1, value=total_dict[key])
            next_row += 1
        wb.save(file_path)
        return 'See ' + filename

    def refresh_cookies(self):
        super(Lowes, self).refresh_cookies()
        self.driver.get('https://www.lowes.com/store/WA-Seattle/0252')
        sleep(2)
        self.driver.find_element_by_id('mainContent').find_elements_by_tag_name('button')[1].click()
        sleep(2)
